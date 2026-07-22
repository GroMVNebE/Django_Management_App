from fnmatch import fnmatch
from django.core.exceptions import ValidationError
from hashids import Hashids
from django.conf import settings
import openpyxl as xl
from decimal import Decimal, ROUND_HALF_UP
import re

hashids = Hashids(salt=settings.SECRET_KEY, min_length=8)


def encode_id(pk: int) -> str:
    """Преобразует числовой id в хэш-строку"""
    return hashids.encode(pk)


def decode_id(hashed_id: str) -> int | None:
    """Преобразует хэш-строку обратно в числовой id"""
    decoded = hashids.decode(hashed_id)
    return decoded[0] if decoded else None


def check_spec(workbook: xl.Workbook):
    """### Проверяет корректность формата спецификации

    Args:
        workbook (xl.Workbook): Excel-книга, содержащая спецификацию

    Raises:
        ValidationError: Причина несоответствия формату

    Returns:
        bool: Соответствует ли спецификация формату
    """
    # Проверка заголовков в спецификации
    required_headers = {
        'Наименование': False,
        'Кол-во': False,
        'ТЗ\nч/д': False,
        'Итого\nруб': False,
        'З/п': False
    }
    sheet = workbook['Спецификация']
    NAME_COLUMN = None
    first_row = next(sheet.iter_rows(
        min_row=1, max_row=1, values_only=False), [])
    for idx, cell in enumerate(first_row):
        val = cell.value
        if val in required_headers:
            required_headers[val] = True
            if val == 'Наименование':
                NAME_COLUMN = idx

    for header, located in required_headers.items():
        if not located:
            raise ValidationError(f'Не найден столбец {header}')

    # Цвет: #33CCFF
    # Закрашено и выделено жирным — заголовок изделия
    # Закрашено без выделения — заголовок части изделия
    rows_generator = sheet.iter_rows(min_row=11, values_only=False)
    # Флаг названия изделия
    isProduct = False
    # Флаг названия части
    isPart = False
    # Флаг наличия изделий
    anyProduct = False

    for row in rows_generator:
        cell = row[NAME_COLUMN]
        name = cell.value

        fill = cell.fill
        is_colored = fill and fill.start_color and fill.start_color.rgb == 'FF33CCFF'

        if is_colored:
            font = cell.font
            is_bold = font and font.bold

            if is_bold:
                if isProduct:
                    raise ValidationError('Обнаружено пустое изделие')
                isProduct = True
                anyProduct = True
            else:
                if isPart:
                    raise ValidationError('Обнаружена пустая часть')
                if not anyProduct:
                    raise ValidationError(
                        'Обнаружена часть, не принадлежащая изделию')
                isPart = True
                isProduct = False

        else:
            if not anyProduct:
                raise ValidationError('Обнаружено оборудование без изделия')
            isProduct = False
            isPart = False

    if not anyProduct:
        raise ValidationError('Не найдено ни одного изделия')

    return True


class ParseableProduct():

    def __init__(self, name: str, price: Decimal, labor_cost: Decimal, payment: Decimal):
        self.name = name
        self.number = ''
        self.price = price
        self.labor_cost = labor_cost
        self.payment = payment
        self.parts: list[ParseablePart] = []
        self.divIntoParts = True


class ParseablePart():

    def __init__(self, name: str, product_price: Decimal, product_payment: Decimal, products: list[ParseableProduct]):
        self.name = name
        self.price = Decimal(0)
        self.payment = Decimal(0)
        self.product_price = product_price
        self.product_payment = product_payment
        for product in products:
            product.parts.append(self)


def parse_spec(spec, blacklist):
    """### Функция для парсинга спецификации
    Парсит изделия и части из спецификации в формате Excel-таблицы для дальнейшего переноса в БД

    Args:
        spec: Загруженная Excel-книга, содержащая спецификацию

    Returns:
        list[`class`:ParseableProduct:]: Список изделий
    """
    # Считываем данные из файла
    spec_format = xl.load_workbook(
        spec, read_only=True, data_only=True)
    sheet = spec_format['Спецификация']
    # Проверяем корректность файла
    check_spec(spec_format)
    # Находим положение нужных столбцов
    required_headers = {
        'Наименование': None,
        'Кол-во': None,
        'ТЗ\nч/д': None,
        'Итого\nруб': None,
        'З/п': None
    }
    first_row = next(sheet.iter_rows(
        min_row=1, max_row=1, values_only=False), [])
    for idx, cell in enumerate(first_row):
        if cell.value in required_headers:
            required_headers[cell.value] = idx
    NAME_COLUMN = required_headers['Наименование']
    AMOUNT_COLUMN = required_headers['Кол-во']
    LABOR_COSTS_COLUMN = required_headers['ТЗ\nч/д']
    PRICE_COLUMN = required_headers['Итого\nруб']
    PAYMENT_COLUMN = required_headers['З/п']

    # Парсим изделия и части
    products: list[ParseableProduct] = []
    parent_products: list[ParseableProduct] = []
    current_part: ParseablePart | None = None
    D2 = Decimal('0.01')
    D4 = Decimal('0.0001')
    rows_generator = sheet.iter_rows(min_row=11, values_only=False)
    for row in rows_generator:
        name_cell = row[NAME_COLUMN]
        name = name_cell.value

        fill = name_cell.fill
        is_colored = fill and fill.start_color and fill.start_color.rgb == 'FF33CCFF'

        amount_val = row[AMOUNT_COLUMN].value
        amount = Decimal(amount_val if amount_val is not None else 1)

        price_val = row[PRICE_COLUMN].value
        price = Decimal(price_val if price_val is not None else 0).quantize(
            D2, rounding=ROUND_HALF_UP)

        # Парсим изделие или часть
        if is_colored:
            font = name_cell.font
            is_bold = font and font.bold

            # Изделие
            if is_bold:
                current_part = None
                parent_products = []

                labor_val = row[LABOR_COSTS_COLUMN].value
                labor_cost = Decimal(labor_val if labor_val is not None else 0).quantize(
                    D4, rounding=ROUND_HALF_UP)

                payment_val = row[PAYMENT_COLUMN].value
                payment = Decimal(payment_val if payment_val is not None else 0).quantize(
                    D2, rounding=ROUND_HALF_UP)

                if amount == 1:
                    products.append(ParseableProduct(
                        name, price, labor_cost, payment))
                    parent_products = [products[-1]]

                elif amount > 1:
                    products_list = name.split(', ')
                    for product_item in products_list:
                        if ' - ' in product_item:
                            start, end = product_item.split(' - ')
                            start_nums = [int(x)
                                          for x in re.findall(r'\d+', start)]
                            end_nums = [int(x)
                                        for x in re.findall(r'\d+', end)]

                            prefix_match = re.match(r'(^[^\d]+)', start)
                            prefix = prefix_match.group(
                                1) if prefix_match else ""

                            results = []
                            if start_nums[:-1] == end_nums[:-1]:
                                stable_nums = start_nums[:-1]
                                full_prefix = prefix + \
                                    ".".join(map(str, stable_nums)) + \
                                    "." if stable_nums else prefix

                                start_last = start_nums[-1]
                                end_last = end_nums[-1]
                                for last in range(start_last, end_last + 1):
                                    results.append(f"{full_prefix}{last}")

                            elif len(start_nums) == 2:
                                for major in range(start_nums[0], end_nums[0] + 1):
                                    s_minor = start_nums[1]
                                    e_minor = end_nums[1]
                                    for minor in range(s_minor, e_minor + 1):
                                        results.append(
                                            f"{prefix}{major}.{minor}")

                            for result in results:
                                products.append(ParseableProduct(
                                    result, price, labor_cost, payment))
                                parent_products.append(products[-1])
                        else:
                            products.append(ParseableProduct(
                                product_item, price, labor_cost, payment))
                            parent_products.append(products[-1])

                if len(parent_products) != amount:
                    raise ValidationError(
                        f'Количество сгенерированных наименований {name} не совпадает с заданным количеством ({len(parent_products)} != {amount})'
                    )

            # Часть
            else:
                current_part = ParseablePart(
                    name, parent_products[0].price, parent_products[0].payment, parent_products)
                if any(fnmatch(name, pattern) for pattern in blacklist):
                    for product in parent_products:
                        product.divIntoParts = False

        # Компонент
        else:
            if current_part:
                current_part.price += price
                current_part.payment = (current_part.product_payment * (
                    current_part.price / current_part.product_price)).quantize(D2, rounding=ROUND_HALF_UP)

    return products
